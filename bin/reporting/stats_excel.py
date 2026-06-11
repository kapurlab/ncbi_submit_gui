"""
Single-labeled-column stats workbook.

vSNP3 stores its run statistics as an ordered ``label -> value`` mapping. Here
we write that mapping as a single labeled column: column A holds the metric
label and column B holds its value, one metric per row — the readable form of
the same data, so an NCBI Submit stats sheet reads like a vSNP3 one.
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Tuple


def write_stats_xlsx(items: List[Tuple[str, str]], path: Path, sample: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "stats"

    header_fill = PatternFill("solid", fgColor="4C8C8A")   # tool teal
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True, color="1F2A2E")
    thin = Side(style="thin", color="E3DED6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Statistic"
    ws["B1"] = "Value"
    for col in ("A1", "B1"):
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].border = border
        ws[col].alignment = Alignment(horizontal="left", vertical="center")

    r = 2
    for label, value in items:
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value="" if value is None else str(value))
        a.font = label_font
        a.border = border
        b.border = border
        a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    max_label = max((len(str(lbl)) for lbl, _ in items), default=20)
    max_value = max((len(str(val)) for _, val in items), default=20)
    ws.column_dimensions["A"].width = min(max(18, max_label + 2), 44)
    ws.column_dimensions["B"].width = min(max(20, max_value + 2), 80)
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
