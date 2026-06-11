"""
Build SRA submission files from FASTQ inputs + canonical metadata.

Per BioProject, writes two tables (TSV and a combined .xlsx) ready for the NCBI
Submission Portal / programmatic submission:
  * <bioproject>_biosample_attributes.tsv — BioSample package attributes
  * <bioproject>_sra_metadata.tsv          — SRA run metadata (one row per run)
and symlinks each sample's FASTQ pair under fastqs/<bioproject>/ so the files
referenced by the metadata are colocated for upload.

Instrument model is read from the FASTQ header when possible, else the preset
default. Fixed library fields come from the preset.
"""

from __future__ import annotations

import gzip
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Instrument inference from the first FASTQ header token. Illumina headers start
# with @<instrument>:<run>:<flowcell>: — the instrument id prefix maps to a model.
_INSTRUMENT_HINTS = [
    (re.compile(r"^@(VH|VL)\d"), "NextSeq 2000"),
    (re.compile(r"^@A\d{5}"), "Illumina NovaSeq 6000"),
    (re.compile(r"^@M\d{5}"), "Illumina MiSeq"),
    (re.compile(r"^@NB\d{6}"), "NextSeq 550"),
    (re.compile(r"^@NS\d{6}"), "NextSeq 500"),
    (re.compile(r"^@K\d{5}"), "Illumina HiSeq 4000"),
    (re.compile(r"^@D\d{5}"), "Illumina HiSeq 2500"),
]

BIOSAMPLE_COLUMNS = [
    "sample_name", "organism", "collection_date", "geo_loc_name",
    "host", "isolation_source", "serotype", "bioproject_accession", "biosample_accession",
]
SRA_COLUMNS = [
    "sample_name", "library_ID", "title", "library_strategy", "library_source",
    "library_selection", "library_layout", "platform", "instrument_model",
    "design_description", "filetype", "filename", "filename2",
]


def find_fastqs(indir: Path, sample_id: str) -> List[Path]:
    """Find the FASTQ file(s) for a sample by prefix match in `indir`."""
    if not indir.is_dir():
        return []
    hits = sorted(
        f for f in indir.glob(f"{sample_id}*.fastq.gz")
        if "_unmapped_" not in f.name
    )
    # Order R1 before R2 when both present.
    hits.sort(key=lambda p: ("_2." in p.name or "_R2" in p.name, p.name))
    return hits


def detect_instrument(fastq: Optional[Path], default: str) -> str:
    if not fastq or not Path(fastq).exists():
        return default
    try:
        with gzip.open(fastq, "rt", errors="replace") as fh:
            header = fh.readline().strip()
    except OSError:
        return default
    for pattern, model in _INSTRUMENT_HINTS:
        if pattern.search(header):
            return model
    return default


def _write_tsv(path: Path, columns: List[str], rows: List[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        fh.write("\t".join(columns) + "\n")
        for r in rows:
            fh.write("\t".join(str(r.get(c, "")) for c in columns) + "\n")


def _write_xlsx(path: Path, sheets: Dict[str, Tuple[List[str], List[Dict[str, str]]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="4C8C8A")
    header_font = Font(bold=True, color="FFFFFF")
    for title, (columns, rows) in sheets.items():
        ws = wb.create_sheet(title[:31])
        ws.append(columns)
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
        for r in rows:
            ws.append([r.get(c, "") for c in columns])
        ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def build_sra(
    samples: List[Dict[str, Any]],
    indir: Path,
    outdir: Path,
    preset: Dict[str, Any],
    date_stamp: str,
    log=print,
) -> Dict[str, Any]:
    """Build per-BioProject SRA files. Returns a summary dict."""
    sra_cfg = preset.get("sra", {}) or {}
    organism = preset.get("organism", "") or ""
    by_bp: Dict[str, Dict[str, List[Dict[str, str]]]] = {}
    missing_files: List[str] = []
    sra_root = outdir / "sra"
    fastq_root = sra_root / "fastqs"

    for s in samples:
        if s.get("excluded") or s.get("sra_existing"):
            continue
        sid = s["sample_id"]
        bp = s.get("bioproject") or "NO_BIOPROJECT"
        fqs = find_fastqs(indir, sid)
        if not fqs:
            missing_files.append(sid)
            log(f"  WARNING: no FASTQ found for {sid} in {indir}")
            continue
        # Colocate the reads for upload.
        dest = fastq_root / bp
        dest.mkdir(parents=True, exist_ok=True)
        linked = []
        for fq in fqs:
            target = dest / fq.name
            if not target.exists():
                try:
                    target.symlink_to(fq.resolve())
                except OSError:
                    pass
            linked.append(fq.name)

        instrument = detect_instrument(fqs[0], sra_cfg.get("instrument_model_default", ""))
        sample_organism = s.get("organism") or organism
        title = sra_cfg.get("title") or f"WGS of {sample_organism}".strip()
        bs_row = {
            "sample_name": sid,
            "organism": sample_organism,
            "collection_date": s.get("collection_date", ""),
            "geo_loc_name": s.get("geo_loc_name", ""),
            "host": s.get("host", ""),
            "isolation_source": s.get("isolation_source", ""),
            "serotype": s.get("serotype", ""),
            "bioproject_accession": s.get("bioproject", ""),
            "biosample_accession": s.get("biosample", ""),
        }
        sra_row = {
            "sample_name": sid,
            "library_ID": sid,
            "title": title,
            "library_strategy": sra_cfg.get("library_strategy", "WGS"),
            "library_source": sra_cfg.get("library_source", "GENOMIC"),
            "library_selection": sra_cfg.get("library_selection", "RANDOM"),
            "library_layout": sra_cfg.get("library_layout", "paired") if len(linked) > 1 else "single",
            "platform": sra_cfg.get("platform", "ILLUMINA"),
            "instrument_model": instrument,
            "design_description": sra_cfg.get("design_description", title),
            "filetype": sra_cfg.get("filetype", "fastq"),
            "filename": linked[0] if linked else "",
            "filename2": linked[1] if len(linked) > 1 else "",
        }
        b = by_bp.setdefault(bp, {"biosample": [], "sra": []})
        b["biosample"].append(bs_row)
        b["sra"].append(sra_row)

    written: List[str] = []
    for bp, tables in by_bp.items():
        bs_tsv = sra_root / f"{bp}_biosample_attributes_{date_stamp}.tsv"
        sra_tsv = sra_root / f"{bp}_sra_metadata_{date_stamp}.tsv"
        xlsx = sra_root / f"{bp}_SRA_{date_stamp}.xlsx"
        _write_tsv(bs_tsv, BIOSAMPLE_COLUMNS, tables["biosample"])
        _write_tsv(sra_tsv, SRA_COLUMNS, tables["sra"])
        _write_xlsx(xlsx, {
            "BioSample": (BIOSAMPLE_COLUMNS, tables["biosample"]),
            "SRA_metadata": (SRA_COLUMNS, tables["sra"]),
        })
        written += [str(bs_tsv), str(sra_tsv), str(xlsx)]
        log(f"  BioProject {bp}: {len(tables['sra'])} run(s) -> {xlsx.name}")

    return {
        "bioprojects": list(by_bp.keys()),
        "n_runs": sum(len(t["sra"]) for t in by_bp.values()),
        "files": written,
        "missing_fastq": missing_files,
        "fastq_root": str(fastq_root),
    }
